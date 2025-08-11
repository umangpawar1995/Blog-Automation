#!/usr/bin/env python3
"""
generate_one_post.py (fixed)

Requirements:
    pip install openpyxl requests pillow

Place linkedin_posting_calendar.xlsx in the same folder.
Set OPENROUTER_API_KEY before running.
"""

import os
import time
import uuid
import base64
import logging
from datetime import datetime

import requests
import openpyxl
from PIL import Image, ImageDraw, ImageFont

# ---------------- CONFIG ----------------
FILE_PATH = "linkedin_posting_calendar.xlsx"
IMAGE_DIR = "images"
OPENROUTER_API_KEY = "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"
TEXT_MODEL = "openai/gpt-3.5-turbo"
IMAGE_MODEL = "stability-ai/stable-diffusion-xl"
MAX_RETRIES = 2
TIMEOUT = 60
HTTP_REFERER = ""  # optional, keep empty unless you need it
WRITE_DEBUG_FILES = False   # set False to avoid creating debug files
# -----------------------------------------

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def ensure_columns(ws):
    headers = [cell.value if cell.value is not None else "" for cell in ws[1]]
    def add_col(name):
        if name not in headers:
            headers.append(name)
            ws.cell(row=1, column=len(headers)).value = name
    add_col("blog")
    add_col("image")
    add_col("status")
    add_col("generated_at")
    headers = [cell.value if cell.value is not None else "" for cell in ws[1]]
    mapping = {}
    for idx, name in enumerate(headers, start=1):
        mapping[name.strip().lower()] = idx
    return mapping


def find_next_row(ws, header_map):
    topic_col = header_map.get("topic", 2)
    status_col = header_map.get("status")
    for r in range(2, ws.max_row + 1):
        topic = ws.cell(row=r, column=topic_col).value
        if not topic:
            continue
        if status_col:
            status = ws.cell(row=r, column=status_col).value
            if status and isinstance(status, str) and status.strip().lower() == "generated":
                continue
        return r
    return None


# ---------- Chat (text) ----------
def call_openrouter_chat(prompt, model=TEXT_MODEL, max_tokens=800, temperature=0.4):
    url = "https://openrouter.ai/api/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
    }
    if HTTP_REFERER:
        headers["HTTP-Referer"] = HTTP_REFERER

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": "You are a professional LinkedIn content writer."},
            {"role": "user", "content": prompt}
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    resp = requests.post(url, headers=headers, json=payload, timeout=TIMEOUT)
    resp.raise_for_status()
    data = resp.json()
    try:
        return data["choices"][0]["message"]["content"]
    except Exception:
        raise RuntimeError(f"Unexpected response structure from OpenRouter chat: {data}")


# ---------- Image helpers ----------
def debug_response(resp, prefix="response"):
    """Return a short text snippet for logging. Optionally write debug file if enabled."""
    try:
        body = resp.text
    except Exception:
        body = "<no-text>"

    snippet = body[:500]
    logging.warning("Image endpoint responded (status=%s). Snippet: %s", resp.status_code, snippet)

    if WRITE_DEBUG_FILES:
        fn = f"debug_{prefix}_{int(time.time())}.txt"
        with open(fn, "w", encoding="utf-8") as f:
            f.write(f"STATUS: {resp.status_code}\n\nHEADERS:\n{resp.headers}\n\nBODY:\n{body}")
        logging.warning("Saved raw response to %s for debugging.", fn)

    return snippet


def download_image_from_url(url, dest_path):
    resp = requests.get(url, stream=True, timeout=TIMEOUT)
    resp.raise_for_status()
    with open(dest_path, "wb") as f:
        for chunk in resp.iter_content(8192):
            f.write(chunk)


def save_b64_image(b64str, dest_path):
    if b64str.startswith("data:"):
        b64str = b64str.split(",", 1)[1]
    img_data = base64.b64decode(b64str)
    with open(dest_path, "wb") as f:
        f.write(img_data)


def text_bbox_size(draw, text, font):
    """
    Return (width, height) of text using best available method across Pillow versions.
    """
    try:
        # Pillow >= 8 has textbbox
        bbox = draw.textbbox((0, 0), text, font=font)
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        return w, h
    except Exception:
        try:
            # older fallback
            w, h = font.getsize(text)
            return w, h
        except Exception:
            # last resort
            return (len(text) * 7, 14)


def generate_local_placeholder_image(topic, dest_path, size=(1200, 630)):
    """
    Robust placeholder hero image creator that doesn't rely on draw.textsize().
    """
    try:
        img = Image.new("RGB", size, color=(18, 33, 79))
        draw = ImageDraw.Draw(img)

        # try to load a common TTF, fallback to default font
        try:
            font = ImageFont.truetype("arial.ttf", 48)
            small_font = ImageFont.truetype("arial.ttf", 20)
        except Exception:
            font = ImageFont.load_default()
            small_font = ImageFont.load_default()

        max_width = size[0] - 120
        words = str(topic).split()
        lines = []
        cur = ""
        for w in words:
            test = (cur + " " + w).strip()
            w_w, _ = text_bbox_size(draw, test, font)
            if w_w <= max_width:
                cur = test
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)

        # vertical centering
        total_h = sum(text_bbox_size(draw, line, font)[1] + 8 for line in lines)
        y = (size[1] - total_h) // 2

        for line in lines:
            w, h = text_bbox_size(draw, line, font)
            x = (size[0] - w) // 2
            draw.text((x, y), line, font=font, fill=(255, 255, 255))
            y += h + 8

        footer = "Data Engineering • AI • Practical Tips"
        fw, fh = text_bbox_size(draw, footer, small_font)
        draw.text(((size[0] - fw) // 2, size[1] - fh - 18), footer, font=small_font, fill=(200, 200, 200))

        img.save(dest_path, format="PNG")
        logging.info("Saved placeholder image to %s", dest_path)
        return dest_path
    except Exception as e:
        logging.exception("Failed to create local placeholder image: %s", e)
        raise


def call_openrouter_image(prompt, model=IMAGE_MODEL, size="1024x1024"):
    """
    Try to generate an image via OpenRouter. Return {"url":...} or {"b64":...}.
    Raises RuntimeError with debug info for non-JSON or unexpected responses.
    """
    url = "https://openrouter.ai/api/v1/images"
    headers = {"Authorization": f"Bearer {OPENROUTER_API_KEY}"}
    if HTTP_REFERER:
        headers["HTTP-Referer"] = HTTP_REFERER

    payload = {"model": model, "prompt": prompt, "size": size, "n": 1}
    resp = requests.post(url, headers=headers, json=payload, timeout=TIMEOUT)

    if resp.status_code != 200:
        snippet = debug_response(resp, prefix="image_error")
        raise RuntimeError(f"Image endpoint returned status {resp.status_code}. Snippet: {snippet}")

    # parse JSON safely
    try:
        data = resp.json()
    except Exception as e:
        snippet = debug_response(resp, prefix="image_nonjson")
        raise RuntimeError(f"Image endpoint returned non-JSON response. Snippet: {snippet}") from e

    # common shapes: {"data":[{...}]}, {"output":[...]}
    d0 = None
    if isinstance(data, dict) and "data" in data and isinstance(data["data"], list) and data["data"]:
        d0 = data["data"][0]
    elif isinstance(data, dict) and "output" in data:
        out = data["output"]
        if isinstance(out, list) and out:
            first = out[0]
            if isinstance(first, str):
                return {"b64": first}
            if isinstance(first, dict):
                d0 = first

    if not d0:
        with open(f"debug_image_json_{int(time.time())}.json", "w", encoding="utf-8") as f:
            import json as _json
            _json.dump(data, f, indent=2)
        raise RuntimeError("Unexpected image JSON structure (debug file saved).")

    if "b64_json" in d0:
        return {"b64": d0["b64_json"]}
    if "b64" in d0:
        return {"b64": d0["b64"]}
    if "url" in d0:
        return {"url": d0["url"]}

    return d0


def make_safe_filename(topic):
    safe = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in topic)[:80].strip()
    unique = uuid.uuid4().hex[:6]
    return f"{safe}_{unique}.png".replace(" ", "_")


# ---------------- Main ----------------
def main():
    if OPENROUTER_API_KEY.startswith("sk-or-v1-REPLACE"):
        logging.error("Please set your OPENROUTER_API_KEY in the script before running.")
        return

    if not os.path.exists(FILE_PATH):
        logging.error("Excel file not found: %s", FILE_PATH)
        return

    os.makedirs(IMAGE_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    header_map = ensure_columns(ws)
    header_map = {k.lower(): v for k, v in header_map.items()}

    # ensure header_map contains needed names
    headers = [ws.cell(row=1, column=c).value or "" for c in range(1, ws.max_column + 1)]
    header_map = {headers[i - 1].strip().lower(): i for i in range(1, len(headers) + 1)}

    next_row = find_next_row(ws, header_map)
    if not next_row:
        logging.info("No rows pending generation.")
        return

    topic = ws.cell(row=next_row, column=header_map.get("topic", 2)).value
    angle = ws.cell(row=next_row, column=header_map.get("angle", 3)).value or ""
    post_format = ws.cell(row=next_row, column=header_map.get("format", 4)).value or ""
    blog_cell = ws.cell(row=next_row, column=header_map.get("blog"))
    image_cell = ws.cell(row=next_row, column=header_map.get("image"))
    status_cell = ws.cell(row=next_row, column=header_map.get("status"))
    generated_at_cell = ws.cell(row=next_row, column=header_map.get("generated_at"))

    logging.info("Processing row %s: %s", next_row, topic)

    try:
        # Generate blog
        if not blog_cell.value or str(blog_cell.value).strip() == "":
            prompt = (
                f"Write a professional, motivational LinkedIn post (~300-400 words) on this topic:\n\n"
                f"Topic: {topic}\n\n"
                f"Angle: {angle}\n\n"
                f"Format: {post_format}\n\n"
                "Tone: storytelling combined with technical expertise. Include a short closing CTA."
            )
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    blog_text = call_openrouter_chat(prompt)
                    if blog_text:
                        blog_cell.value = blog_text.strip()
                        logging.info("Blog generated (len=%s chars)", len(blog_text))
                        break
                except Exception as e:
                    logging.warning("Attempt %s: blog generation failed: %s", attempt, e)
                    if attempt == MAX_RETRIES:
                        raise

        # Generate image (or fallback)
        if not image_cell.value or str(image_cell.value).strip() == "":
            image_prompt = f"Create a clean, professional LinkedIn hero image for: '{topic}'. Theme: {angle}. Style: modern, minimal, high-contrast, suitable for a technical audience."
            img_resp = None
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    img_resp = call_openrouter_image(image_prompt)
                    break
                except Exception as e:
                    logging.warning("Attempt %s: image generation failed: %s", attempt, e)
                    # if last attempt, we'll fallback below
                    if attempt == MAX_RETRIES:
                        img_resp = None

            filename = make_safe_filename(topic)
            local_path = os.path.join(IMAGE_DIR, filename)

            if img_resp:
                try:
                    if "url" in img_resp:
                        download_image_from_url(img_resp["url"], local_path)
                        logging.info("Image downloaded to %s", local_path)
                    elif "b64" in img_resp:
                        save_b64_image(img_resp["b64"], local_path)
                        logging.info("Image saved (decoded base64) to %s", local_path)
                    else:
                        logging.warning("Unexpected image response, creating placeholder.")
                        generate_local_placeholder_image(topic, local_path)
                except Exception as e:
                    logging.exception("Failed to save/generated image: %s", e)
                    generate_local_placeholder_image(topic, local_path)
            else:
                # final fallback - create placeholder
                generate_local_placeholder_image(topic, local_path)

            image_cell.value = local_path

        # mark status & timestamp
        status_cell.value = "generated"
        generated_at_cell.value = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

        # save workbook
        try:
            wb.save(FILE_PATH)
            logging.info("Saved updates to workbook: %s", FILE_PATH)
        except PermissionError:
            alt = f"{os.path.splitext(FILE_PATH)[0]}_updated_{int(time.time())}.xlsx"
            wb.save(alt)
            logging.warning("Original file locked. Saved to alternate file: %s", alt)

    except Exception as e:
        logging.exception("Failed to process row %s: %s", next_row, e)
        try:
            status_cell.value = f"error: {str(e)[:200]}"
            wb.save(FILE_PATH)
        except Exception:
            logging.exception("Failed to write error status to disk.")
        return

    logging.info("Done for row %s", next_row)


if __name__ == "__main__":
    main()
