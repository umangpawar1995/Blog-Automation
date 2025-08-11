import openpyxl
import requests

# ==== CONFIG ====
OPENROUTER_API_KEY = "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"
MODEL = "openai/gpt-3.5-turbo"
FILE_PATH = "blog_topics.xlsx"

# ==== LOAD EXCEL ====
wb = openpyxl.load_workbook(FILE_PATH)
sheet = wb["Sheet1"]

# ==== PROCESS TOPICS ====
for row in range(2, sheet.max_row + 1):  # skip header
    topic = sheet.cell(row=row, column=1).value
    blog_text = sheet.cell(row=row, column=2).value

    if topic and not blog_text:
        print(f"Generating blog for: {topic}")
        payload = {
            "model": MODEL,
            "messages": [
                {"role": "system", "content": "You are a professional blog writer."},
                {"role": "user", "content": f"Write a 300-word blog on the topic: {topic}"}
            ],
            "temperature": 0.7,
            "max_tokens": 500
        }
        headers = {
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://yourdomain.com",
            "X-Title": "Local Excel Blog Automation"
        }

        response = requests.post("https://openrouter.ai/api/v1/chat/completions",
                                 headers=headers, json=payload)

        if response.status_code == 200:
            blog = response.json()['choices'][0]['message']['content']
            sheet.cell(row=row, column=2).value = blog
            print(f"✅ Blog for '{topic}' saved to Excel.")
        else:
            print(f"❌ Error for '{topic}': {response.text}")

# ==== SAVE CHANGES ====
wb.save(FILE_PATH)
print("All blogs generated and saved!")
